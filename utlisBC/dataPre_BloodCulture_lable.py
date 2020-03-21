'''
Created on 2019年6月9日

标记血培养结果 
阳性患者：同一个患者 一次血培养结果为阳性 标注为1 只抽取第一次血培养时间 
阴性患者：同一个患者 一次血培养结果为阴性 标注为0 检验时间不同的多次阴性 均标注为0 检验时间相同的 只标注1次

OUT: 阳性文件：BCu--POS.xlsx
阴性文件：BCu--NEG.xlsx

@author: cmingwhu
'''

from openpyxl import load_workbook ,Workbook
import os,sys
import time as tm
import numpy as np
import argparse

np.random.seed(2018)

# 默认处理阳性数据
# BCu_flag = True  阳性
#BCu_flag = False  阴性
# BCu_flag = True

# HOSP_ID: 住院号  TEST_NUM：检验号   TEST_TIME：检验时间   SEX：性别
BCu_column = {'病人ID':None,'检验唯一号':None,'检验日期':None,'病人性别':None,'病人年龄':None,'申请科室名称':None,
              '诊断':None,'检验目的名称':None,'检验结果':None,'结果标志':None}

BCu_dic = {'病人ID':None,'检验唯一号':None,'检验日期':None,'病人性别':None,'病人年龄':None,'检验结果':None}

BCu_firRow = ['HOSP_ID','TEST_NUM','TEST_TIME','SEX','AGE','FLAG']

BCu_row = []

BCu_NEG_flag = '无细菌生长'
BCu_NEG_flag_two = '未见'


output_file = {'BCu_file': None, 'error_file': None}
output_wb = {'BCu_wb': None, 'error_wb': None}
output_sheet = {'BCu_sheet': None, 'error_sheet': None}

error_debug = False
info_dedug = False

def initDic():
    for key in BCu_dic.keys():
        BCu_dic[key] = None

'''
'''
def initOutFile(path,BCu_PN_flag):
    tm.sleep(2)
    cur_time = (tm.strftime('%Y-%m-%d %H:%M:%S', tm.localtime(tm.time()))).replace(':', '-', 100)

    if BCu_PN_flag == 'True': 
        output_file['BCu_file'] = path + '--BCu ' + '(' + cur_time + ')' + '.xlsx'
        output_file['error_file'] = path + '--error ' + '(' + cur_time + ')' + '.xlsx'
    else:
        output_file['BCu_file'] = path + '--BCu ' + '(' + cur_time + ')' + '.xlsx'
        output_file['error_file'] = path + '--error ' + '(' + cur_time + ')' + '.xlsx'

    # 添加第一行
    for key in BCu_dic.keys():
        BCu_dic[key] = key
        
    for key in output_file.keys():
        file = output_file[key]
        if os.path.exists(file):
            os.remove(output_file[key])
            print('file is the same,return\n')
            exit()
        else:            
            row =[]
            wb = Workbook()
            sheet = wb.active          
            sheet.append(BCu_firRow)                  
            wb.save(file) 

    for key in output_wb.keys():
        file = output_file[key.rstrip('wb') + ('file')]
        key_sheet = key.rstrip('wb') + ('sheet')
        output_wb[key] = load_workbook(file)
        output_sheet[key_sheet] = output_wb[key].get_active_sheet()


#     writeFile(output_file['BCu_file'], output_wb['BCu_wb'], output_sheet['BCu_sheet'], **BCu_dic)

'''
'''
def writeFile(wr_file, wb, sheet, **dic):
    row = []
    err_flag = False

    for key,value in dic.items():
        if(value == None):
            if(error_debug == False):
                return
            #print('非法的数据 %s \n'%key)
            err_flag = True
        row.append(value)
    BCu_row.append(row)

    if(err_flag == True):
        wr_excel = output_file['error_file']
        wb = output_wb['error_wb']
        sheet = output_sheet['error_sheet']
    else:
        wr_excel = wr_file

    sheet.append(row)
#     wb.save(wr_excel) # 浪费时间的语句

'''
血培养阳性
'''
def getPOSInf(dataFile):
    csv = load_workbook(dataFile)
    # 获取sheet：
    sheet = csv.get_active_sheet()
#     rows = sheet.max_row     # 获取行数，浪费时间多
#     cols = sheet.max_column  # 获取列数，浪费时间多
#     print('rows : ',rows)
#     print('column : ',cols)

    # 取出列名所在的index
    tempList  = list(sheet.rows)
    for key in BCu_column.keys():
        column_no = 1
        for cell in tempList[0]:
            if(cell.value == key):
                BCu_column[key] = column_no
                break
            column_no += 1
            
            
    # 获取检验号一列的数据
    rows_no = 2
    # 标志字段
    BCu_flag = False
    write_flag = False

    jianyanhao_cell_value_old = 'None'
    jianyanhao_cell_value = 'None'
    jianyanTime_cell_value_old = 'None'
    jianyanTime_cell_value = 'None'
    zhuyuanID_cell_value = 'None'
    zhuyuanID_cell_value_old ='None'
    
    jianyanhao_index = BCu_column['检验唯一号']
    jianyanmudiname_index = BCu_column['检验目的名称']
    jieguobiaozhi_index = BCu_column['结果标志']
    jianyanTime_index = BCu_column['检验日期']
    zhuyuanID_index = BCu_column['病人ID']
    
    if (info_dedug):
        begin_time = tm.clock()
        
    BCu_HOSP_ID = None
    for cell in list(sheet.columns)[jianyanhao_index]:      
        # 抽取检验目的名称
        jianyanmudiname_cell = sheet.cell(row=rows_no, column=jianyanmudiname_index)
        jianyanmudiname_cell_value = str(jianyanmudiname_cell.value)
        
        
        if ('血培养' in jianyanmudiname_cell_value):
            # 抽取检验号
            jianyanhao_cell = sheet.cell(row=rows_no, column=jianyanhao_index)
            jianyanhao_cell_value = str(jianyanhao_cell.value)    
            # 抽取检验时间
            jianyanTime_cell = sheet.cell(row=rows_no, column=jianyanTime_index)
            jianyanTime_cell_value = str(jianyanTime_cell.value)
            # 抽取住院号
            zhuyuanID_cell = sheet.cell(row=rows_no, column=zhuyuanID_index)
            zhuyuanID_cell_value = str(zhuyuanID_cell.value)
          

            # 抽取结果标志
            jieguobiaozhi_cell = sheet.cell(row=rows_no, column=jieguobiaozhi_index)
            jieguobiaozhi_cell_value = str(jieguobiaozhi_cell.value)
            # 抽取检验结果
            jianyanjieguo_cell = sheet.cell(row=rows_no, column=BCu_column['检验结果'])
            jianyanjieguo_cell_value = str(jianyanjieguo_cell.value)

            # 检验号的变化，作为上一批写入标志
            if(zhuyuanID_cell_value != zhuyuanID_cell_value_old):
                write_flag = True
            else:
                if(jianyanhao_cell_value != jianyanhao_cell_value_old) and (jianyanTime_cell_value != jianyanTime_cell_value_old):
                    write_flag = True

            # 遍历以前的值，过滤重复：a，b，a-> a，b
            #BCu_row_revs = list(reversed(BCu_row))
            if write_flag == True:
                for i in range(0, len(BCu_row)):
                    if BCu_dic['病人ID'] == BCu_row[i][0] and BCu_dic['检验日期'] == BCu_row[i][2]:
                        write_flag = False
                        break

            if write_flag == True:
                if (BCu_flag == True):
                    writeFile(output_file['BCu_file'], output_wb['BCu_wb'], output_sheet['BCu_sheet'],
                              **BCu_dic)
                    BCu_HOSP_ID = BCu_dic['病人ID']
                    BCu_flag = False                   
                write_flag = False
                initDic()

            # 阳性患者
            if ((jieguobiaozhi_cell_value != 'None') or \
                ((BCu_NEG_flag in jianyanjieguo_cell_value) == False and \
                 ((BCu_NEG_flag_two in jianyanjieguo_cell_value) == False))):
                if((BCu_NEG_flag_two in jieguobiaozhi_cell_value) == False):
                    BCu_flag = True
                    for key in BCu_dic.keys():
                        if key == '检验结果':
                            BCu_dic[key] = '1'
                        else:
                            BCu_dic[key] = str(sheet.cell(row=rows_no, column=BCu_column[key]).value)
                    if BCu_HOSP_ID == BCu_dic['病人ID']:
                        BCu_flag = False
            else:  # 阴性患者
                pass

            jianyanhao_cell_value_old = jianyanhao_cell_value
            jianyanTime_cell_value_old = jianyanTime_cell_value
            zhuyuanID_cell_value_old = zhuyuanID_cell_value
            rows_no += 1
        else:
            rows_no += 1
                
        # times
        if (info_dedug and ((rows_no % 5000) == 0)):
            print(rows_no)
            print('spend time (second): ',tm.clock()- begin_time)
            begin_time = tm.clock()

    for i in range(0, len(BCu_row)):
        if BCu_dic['病人ID'] == BCu_row[i][0] and BCu_dic['检验日期'] == BCu_row[i][2]:
            BCu_flag = False
            break
    if (BCu_flag == True):
        writeFile(output_file['BCu_file'], output_wb['BCu_wb'], output_sheet['BCu_sheet'],
                  **BCu_dic)
    print(rows_no)


 
'''
血培养阴性
'''
def getNEGInf(dataFile):
    csv = load_workbook(dataFile)
    # 获取sheet：
    sheet = csv.get_active_sheet()

    # 取出列名所在的index
    tempList = list(sheet.rows)
    for key in BCu_column.keys():
        column_no = 1
        for cell in tempList[0]:
            if (cell.value == key):
                BCu_column[key] = column_no
                break
            column_no += 1

    # 获取检验号一列的数据
    rows_no = 2
    # 标志字段
    BCu_flag = False
    write_flag = False

    jianyanhao_cell_value_old = 'None'
    jianyanhao_cell_value = 'None'
    jianyanTime_cell_value_old = 'None'
    jianyanTime_cell_value = 'None'
    zhuyuanID_cell_value = 'None'
    zhuyuanID_cell_value_old = 'None'

    jianyanhao_index = BCu_column['检验唯一号']
    jianyanmudiname_index = BCu_column['检验目的名称']
    jieguobiaozhi_index = BCu_column['结果标志']
    jianyanTime_index = BCu_column['检验日期']
    zhuyuanID_index = BCu_column['病人ID']

    if (info_dedug):
        begin_time = tm.clock()

    BCu_HOSP_ID = None
    for cell in list(sheet.columns)[jianyanhao_index]:
        # 抽取检验目的名称
        jianyanmudiname_cell = sheet.cell(row=rows_no, column=jianyanmudiname_index)
        jianyanmudiname_cell_value = str(jianyanmudiname_cell.value)

        if ('血培养' in jianyanmudiname_cell_value):
            # 抽取检验号
            jianyanhao_cell = sheet.cell(row=rows_no, column=jianyanhao_index)
            jianyanhao_cell_value = str(jianyanhao_cell.value)
            # 抽取检验时间
            jianyanTime_cell = sheet.cell(row=rows_no, column=jianyanTime_index)
            jianyanTime_cell_value = str(jianyanTime_cell.value)
            # 抽取住院号
            zhuyuanID_cell = sheet.cell(row=rows_no, column=zhuyuanID_index)
            zhuyuanID_cell_value = str(zhuyuanID_cell.value)

            # 抽取结果标志
            jieguobiaozhi_cell = sheet.cell(row=rows_no, column=jieguobiaozhi_index)
            jieguobiaozhi_cell_value = str(jieguobiaozhi_cell.value)
            # 抽取检验结果
            jianyanjieguo_cell = sheet.cell(row=rows_no, column=BCu_column['检验结果'])
            jianyanjieguo_cell_value = str(jianyanjieguo_cell.value)
                    
            # 检验号的变化，作为上一批写入标志
            if (zhuyuanID_cell_value != zhuyuanID_cell_value_old):
                write_flag = True
            else:
                if (jianyanhao_cell_value != jianyanhao_cell_value_old) and (jianyanTime_cell_value != jianyanTime_cell_value_old):
                    write_flag = True
            
            
            # 遍历以前的值，过滤重复：a，b，a-> a，b
            # BCu_row_revs = list(reversed(BCu_row))
            if write_flag == True:
                for i in range(0, len(BCu_row)):
                    if BCu_dic['病人ID'] == BCu_row[i][0] and BCu_dic['检验日期'] == BCu_row[i][2]:
                        write_flag = False
                        break

            if write_flag == True:
                if (BCu_flag == True):
                    writeFile(output_file['BCu_file'], output_wb['BCu_wb'], output_sheet['BCu_sheet'],
                              **BCu_dic)
                    BCu_HOSP_ID = BCu_dic['病人ID']
                    BCu_flag = False
                write_flag = False
                initDic()
                
            # 阴性患者
            if ((jieguobiaozhi_cell_value != 'None') or \
                    ((BCu_NEG_flag in jianyanjieguo_cell_value) == False and \
                     ((BCu_NEG_flag_two in jianyanjieguo_cell_value) == False))):
                if ((BCu_NEG_flag_two in jieguobiaozhi_cell_value) == False):
                    pass # 阳性患者
                else:# 阴性患者
                    BCu_flag = True
                    for key in BCu_dic.keys():
                        if key == '检验结果':
                            BCu_dic[key] = '0'
                        else:
                            BCu_dic[key] = str(sheet.cell(row=rows_no, column=BCu_column[key]).value)
                    if BCu_HOSP_ID == BCu_dic['病人ID']:
                        BCu_flag = False
            else:  # 阴性患者
                BCu_flag = True
                for key in BCu_dic.keys():
                    if key == '检验结果':
                        BCu_dic[key] = '0'
                    else:
                        BCu_dic[key] = str(sheet.cell(row=rows_no, column=BCu_column[key]).value)
                if BCu_HOSP_ID == BCu_dic['病人ID']:
                    BCu_flag = False

            jianyanhao_cell_value_old = jianyanhao_cell_value
            jianyanTime_cell_value_old = jianyanTime_cell_value
            zhuyuanID_cell_value_old = zhuyuanID_cell_value
            rows_no += 1
        else:
            rows_no += 1

        # times
        if (info_dedug and ((rows_no % 5000) == 0)):
            print(rows_no)
            print('spend time (second): ', tm.clock() - begin_time)
            begin_time = tm.clock()

    for i in range(0, len(BCu_row)):
        if BCu_dic['病人ID'] == BCu_row[i][0] and BCu_dic['检验日期'] == BCu_row[i][2]:
            BCu_flag = False
            break
    if (BCu_flag == True):
        writeFile(output_file['BCu_file'], output_wb['BCu_wb'], output_sheet['BCu_sheet'],
                  **BCu_dic)
    print(rows_no)


def _process_args():
    parser = argparse.ArgumentParser(description='This is a temporal expression extraction sample program')
    parser.add_argument('-i', default="None", help='input file path')
    parser.add_argument('-f', default=True, help='POS: True, NEG: False')

    return parser.parse_args(sys.argv[1:])



if __name__ == '__main__':
    
    main_time = tm.clock()
 
    # 命名行
    parser = argparse.ArgumentParser()
    parser.add_argument('-i', default="None", help='input file')
    parser.add_argument('-f', default='True', help='POS: True, NEG: False')
    args = parser.parse_args()
    
    
    file_name = args.i
    BCu_PN_flag = args.f  
    
    if file_name == 'None':
         print('缺少参数：文件名, 默认处理阳性文件 ！eg: dataPre_LisItem_Segmentation -i file -f True (默认)')
         sys.exit()

#     file_name = 'XPY01NEG.xlsx'
    file_dir = os.path.dirname(os.path.realpath(__file__))
 
    if sys.platform == 'linux':
        input_file = file_dir + r'/' + file_name
    else:
#         file_dir = file_dir.rstrip(r'\utlisBC')    
#         input_file = file_dir + str(r'\?').rstrip(r'?')+'data'+ str(r'\?').rstrip(r'?')+'BC'+ str(r'\?').rstrip(r'?')+file_name
        input_file = os.path.join(os.path.abspath('..'), 'data', 'BC',file_name)  

    file_path = input_file.rstrip('.xlsx')
 
    print(sys.platform)
    print(file_name)
    print(file_dir)
    print(file_path)
    print(input_file)
 
    initDic()
    initOutFile(file_path,BCu_PN_flag)
 
    # 调试
    info_dedug = True
    error_debug = False

    if BCu_PN_flag == 'True':
        getPOSInf(input_file)
    else:
        getNEGInf(input_file)

    # 保存文件
    for key in output_file.keys():
        key_wb = key.rstrip('file') + ('wb')
        wb = output_wb[key_wb]
        wb.save(output_file[key])

    print('spend total time (second): ', tm.clock() - main_time)