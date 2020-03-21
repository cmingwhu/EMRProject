'''
Created on 2019年6月8日

分割 LIS 项目：血常规BRe  降钙素原PCT  C反应蛋白CRP 血凝实验HA

@author: cmingwhu
'''
from openpyxl import load_workbook ,Workbook
import os,sys
import time as tm
import numpy as np
import argparse

np.random.seed(2018)


# 血常规项目 Blood routine(BRe):
# 白细胞计数     红细胞计数     血红蛋白    血小板计数     中性粒细胞百分数     淋巴细胞百分数    单核细胞百分数
# 嗜酸性粒细胞百分数  嗜碱性粒细胞百分数     中性粒细胞绝对值     淋巴细胞绝对值    单核细胞绝对值    嗜酸性粒细胞绝对值
# 嗜碱性粒细胞绝对值    红细胞压积    平均红细胞体积     平均红细胞血红蛋白含量     平均红细胞血红蛋白浓度    红细胞分布宽度
# 平均血小板体积     血小板压积    血小板分布宽度    有核红细胞计数    有核红细胞百分比
# HOSP_ID: 住院号  TEST_NUM：检验号   TEST_TIME：检验时间   SEX：性别
BRe_xmdm = ['WBC', 'RBC', 'Hb', 'PLT', 'Neut%', 'Lymph%', 'Mono%', 'Eos%', 'Baso%', 'Neut#', 'Lymph#',
            'Mono#', 'Eos#', 'Baso#', 'Hct', 'MCV', 'MCH', 'MCHC', 'RDW', 'MPV', 'Pct', 'PDW', 'NRBC', 'NRBC%']
BRe_dic = {'HOSP_ID': None, 'TEST_NUM': None,'TEST_TIME': None, 'SEX': None, 
           'WBC': None, 'RBC': None, 'Hb': None, 'PLT': None, 'Neut%': None, 'Lymph%': None,
           'Mono%': None,  'Eos%': None, 'Baso%': None, 'Neut#': None, 'Lymph#': None, 'Mono#': None, 'Eos#': None,
           'Baso#': None, 'Hct': None, 'MCV': None, 'MCH': None, 'MCHC': None, 'RDW': None, 'MPV': None, 'Pct': None,
           'PDW': None, 'NRBC': None, 'NRBC%': None}

# 血凝实验项目 HA:
# 凝血酶原时间    凝血酶原时间活动度    国际标准化比值    活化部分凝血活酶时间    纤维蛋白原测定    凝血酶时间
# D-二聚体    纤维蛋白（原）降解产物
HA_xmdm = ['PT', 'INR', 'PT-%','PT%','APTT', 'FIB','Fib', 'TT', 'D-Dimer', 'FDP']
HA_dic = {'HOSP_ID': None, 'TEST_NUM': None,'TEST_TIME': None, 'SEX': None, 
            'PT':None, 'PT-%':None, 'INR':None, 'APTT':None, 'FIB':None, 'TT':None, 'D-Dimer':None, 'FDP':None}

# 降钙素原 PCT:
PCT_xmdm = ['PCT']
PCT_dic = {'HOSP_ID': None, 'TEST_NUM': None,'TEST_TIME': None, 'SEX': None, 'PCT': None}
PCT_xmmc = '降钙素原'

# C反应蛋白 CRP: 
CRP_xmdm = ['CRP','hsCRP']
CRP_dic = {'HOSP_ID': None, 'TEST_NUM': None,'TEST_TIME': None, 'SEX': None, 'CRP': None}
CRP_xmmc = '反应蛋白'

# input_file = r'D:\ZXL\pycharm\project\u2\201801-1.xlsx'
# input_file = .\data\BC\201801.xlsx
output_file = {'BRe_file': None, 'HA_file': None, 'PCT_file': None , 'CRP_file': None ,'error_file': None}
output_wb = {'BRe_wb': None, 'HA_wb': None, 'PCT_wb': None , 'CRP_wb': None ,'error_wb': None}
output_sheet = {'BRe_sheet': None, 'HA_sheet': None, 'PCT_sheet': None , 'CRP_sheet': None ,'error_sheet': None}

error_debug = False
info_dedug = False

def initDic():
    for key in BRe_dic.keys():
        BRe_dic[key] = None
    for key in HA_dic.keys():
        HA_dic[key] = None
    for key in PCT_dic.keys():
        PCT_dic[key] = None
    for key in CRP_dic.keys():
        CRP_dic[key] = None

    
def initOutFile(path):
    tm.sleep(2)
    cur_time = (tm.strftime('%Y-%m-%d %H:%M:%S', tm.localtime(tm.time()))).replace(':', '-', 100)

    # 添加第一行
    for key in BRe_dic.keys():
        BRe_dic[key] = key
    for key in HA_dic.keys():
        HA_dic[key] = key
    for key in PCT_dic.keys():
        PCT_dic[key] = key
    for key in CRP_dic.keys():
        CRP_dic[key] = key

    output_file['BRe_file'] = path + '--BRe ' + '(' + cur_time + ')' + '.xlsx'
    output_file['HA_file'] = path + '--HA ' + '(' + cur_time + ')' + '.xlsx'
    output_file['PCT_file'] = path + '--PCT ' + '(' + cur_time + ')' + '.xlsx'
    output_file['CRP_file'] = path + '--CRP ' + '(' + cur_time + ')' + '.xlsx'
    output_file['error_file'] = path + ' --error ' + '(' + cur_time + ')' + '.xlsx'

    for key in output_file.keys():
        file = output_file[key]
        if os.path.exists(file):
            os.remove(output_file[key])
            print('file is the same,return\n')
            exit()
        else:
            row =[]
            wb = Workbook()
            sheet = wb.active
            if (key.rstrip('file') == 'BRe_'):
                for key,value in BRe_dic.items():
                   row.append(value)
                sheet.append(row)
            elif (key.rstrip('file') == 'HA_'):
                for key,value in HA_dic.items():
                   row.append(value)
                sheet.append(row)
            elif (key.rstrip('file') == 'PCT_'):
                for key,value in PCT_dic.items():
                   row.append(value)
                sheet.append(row)
            elif (key.rstrip('file') == 'CRP_'):
                for key,value in CRP_dic.items():
                   row.append(value)
                sheet.append(row)                  
            wb.save(file) 
            
                       
    for key in output_wb.keys():
        file = output_file[key.rstrip('wb') + ('file')]
        key_sheet = key.rstrip('wb') + ('sheet')
        output_wb[key] = load_workbook(file)
        output_sheet[key_sheet] = output_wb[key].get_active_sheet()
 
#     writeFile(output_file['BRe_file'], output_wb['BRe_wb'], output_sheet['BRe_sheet'], **BRe_dic)
#     writeFile(output_file['HA_file'], output_wb['HA_wb'], output_sheet['HA_sheet'], **HA_dic)
#     writeFile(output_file['PCT_file'], output_wb['PCT_wb'], output_sheet['PCT_sheet'], **PCT_dic)
#     writeFile(output_file['CRP_file'], output_wb['CRP_wb'], output_sheet['CRP_sheet'], **CRP_dic)

def writeFile(wr_file, wb, sheet, **dic):
    row = []
    err_flag = False

    for key,value in dic.items():
        if(value == None):
            if(error_debug == False):
                return
            #print('非法的数据 %s \n'%key)
            err_flag = True
        row.append(value)

    if(err_flag == True):
        wr_excel = output_file['error_file']
        wb = output_wb['error_wb']
        sheet = output_sheet['error_sheet']
    else:
        wr_excel = wr_file

    sheet.append(row)
#     wb.save(wr_excel) # 浪费时间的语句

# LIS项目分割
def getLisInf(dataFile):
    csv = load_workbook(dataFile)
    # 获取sheet：
    sheet = csv.get_active_sheet()
#     rows = sheet.max_row     # 获取行数，浪费时间多
#     cols = sheet.max_column  # 获取列数，浪费时间多
#     print('rows : ',rows)
#     print('column : ',cols)

    column_no = 1
    zhuyuanhao_index = None
    sex_index = None
    jianyanhao_index = None
    xiangmudaima_index = None
    xiangmumingcheng_index = None
    jianyanshijian_index = None
    jianyanjieguo_index = None

    # 取出列名所在的index
    for cell in list(sheet.rows)[0]:
        if(cell.value == '病人ID'):      
            zhuyuanhao_index = column_no
        if(cell.value == '病人性别'):     
            sex_index = column_no
        if(cell.value == '检验唯一号'):   
            jianyanhao_index = column_no
        if(cell.value == '项目代号'):   
            xiangmudaima_index = column_no
        if(cell.value == '项目名称'):   
            xiangmumingcheng_index = column_no
        if(cell.value == '检验结果'):
            jianyanjieguo_index = column_no
        if(cell.value == '检验日期'):
            jianyanshijian_index = column_no

        column_no += 1

    # 获取检验号一列的数据
    rows_no = 2

    jianyanhao_cell_value_old = 'None'
    jianyanhao_cell_value = 'None'

    # 标志字段
    BRe_flag = False
    HA_flag = False
    PCT_flag = False
    CRP_flag = False

    write_flag = False

    begin_time = tm.clock()

    for cell in list(sheet.columns)[jianyanhao_index]:
        jianyanhao_cell = sheet.cell(row=rows_no, column=jianyanhao_index)
        jianyanhao_cell_value = str(jianyanhao_cell.value)
           
        # 相同的检验号，作为同一批处理
        if(jianyanhao_cell_value == jianyanhao_cell_value_old) or (rows_no == 2):
            xiangmudaima_cell = sheet.cell(row=rows_no, column=xiangmudaima_index)
            xiangmudaima_cell_value = str(xiangmudaima_cell.value)

            # 项目名称，作为降钙素原和C反应蛋白判断的联合项
            # xiangmumingcheng_cell = sheet.cell(row=rows_no, column=xiangmumingcheng_index)
            # xiangmumingcheng_cell_value = str(xiangmumingcheng_cell.value)

            # 取出检验结果
            if(xiangmudaima_cell_value in BRe_xmdm != 0):
                BRe_flag = True
                BRe_dic[xiangmudaima_cell_value] = str(sheet.cell(row=rows_no, column=jianyanjieguo_index).value)
                if((BRe_dic['HOSP_ID']) == None) :
                    BRe_dic['HOSP_ID'] = str(sheet.cell(row=rows_no, column=zhuyuanhao_index).value)
                    BRe_dic['TEST_NUM'] = str(sheet.cell(row=rows_no, column=jianyanhao_index).value)
                    BRe_dic['SEX'] = str(sheet.cell(row=rows_no, column=sex_index).value)
                    BRe_dic['TEST_TIME'] = str(sheet.cell(row=rows_no, column=jianyanshijian_index).value)
            elif(xiangmudaima_cell_value in HA_xmdm):
                HA_flag = True
                if xiangmudaima_cell_value =='PT%':
                    xiangmudaima_cell_value = 'PT-%'
                if xiangmudaima_cell_value == 'Fib':
                    xiangmudaima_cell_value = 'FIB'              
                HA_dic[xiangmudaima_cell_value] = str(sheet.cell(row=rows_no, column=jianyanjieguo_index).value)
                if((HA_dic['HOSP_ID']) == None) :
                    HA_dic['HOSP_ID'] = str(sheet.cell(row=rows_no, column=zhuyuanhao_index).value)
                    HA_dic['TEST_NUM'] = str(sheet.cell(row=rows_no, column=jianyanhao_index).value)
                    HA_dic['SEX'] = str(sheet.cell(row=rows_no, column=sex_index).value)
                    HA_dic['TEST_TIME'] = str(sheet.cell(row=rows_no, column=jianyanshijian_index).value)
            # elif(xiangmudaima_cell_value in PCT_xmdm) and (PCT_xmmc in xiangmumingcheng_cell_value):
            elif (xiangmudaima_cell_value in PCT_xmdm):
                xiangmumingcheng_cell = sheet.cell(row=rows_no, column=xiangmumingcheng_index)
                xiangmumingcheng_cell_value = str(xiangmumingcheng_cell.value)
                if(PCT_xmmc in xiangmumingcheng_cell_value):
                    # 更新写标志位，是一个单项项目，只要有一次就表示已经可以写入
                    write_flag = True
                    PCT_flag = True
                    PCT_dic[xiangmudaima_cell_value] = str(sheet.cell(row=rows_no, column=jianyanjieguo_index).value)
                    if((PCT_dic['HOSP_ID']) == None) :
                        PCT_dic['HOSP_ID'] = str(sheet.cell(row=rows_no, column=zhuyuanhao_index).value)
                        PCT_dic['TEST_NUM'] = str(sheet.cell(row=rows_no, column=jianyanhao_index).value)
                        PCT_dic['SEX'] = str(sheet.cell(row=rows_no, column=sex_index).value)
                        PCT_dic['TEST_TIME'] = str(sheet.cell(row=rows_no, column=jianyanshijian_index).value)
            # elif(xiangmudaima_cell_value in CRP_xmdm) and (CRP_xmmc in xiangmumingcheng_cell_value):
            elif (xiangmudaima_cell_value in CRP_xmdm):
                # 项目名称，作为降钙素原和C反应蛋白判断的联合项
                xiangmumingcheng_cell = sheet.cell(row=rows_no, column=xiangmumingcheng_index)
                xiangmumingcheng_cell_value = str(xiangmumingcheng_cell.value)
                if(CRP_xmmc in xiangmumingcheng_cell_value):
                    # 更新写标志位，CRP存在同一个检验号下有hsCRP和CRP，且是一个单项项目，只要有一次就表示已经可以写入
                    write_flag = True
                    CRP_flag = True
                    if xiangmudaima_cell_value == 'hsCRP':
                        xiangmudaima_cell_value = 'CRP'
                    CRP_dic[xiangmudaima_cell_value] = str(sheet.cell(row=rows_no, column=jianyanjieguo_index).value)
                    if((CRP_dic['HOSP_ID']) == None) :
                        CRP_dic['HOSP_ID'] = str(sheet.cell(row=rows_no, column=zhuyuanhao_index).value)
                        CRP_dic['TEST_NUM'] = str(sheet.cell(row=rows_no, column=jianyanhao_index).value)
                        CRP_dic['SEX'] = str(sheet.cell(row=rows_no, column=sex_index).value)
                        CRP_dic['TEST_TIME'] = str(sheet.cell(row=rows_no, column=jianyanshijian_index).value)
            else:
                # 其余不做处理
                pass
            '''
            # 更新写标志位
            if(rows_no  == rows):
                write_flag = True
            '''
        else:
             # 如果检验号不同，则写文件
             # 更新写标志位
             write_flag = True

        if(write_flag == True):
            if(BRe_flag == True):
                writeFile(output_file['BRe_file'], output_wb['BRe_wb'], output_sheet['BRe_sheet'], **BRe_dic)
                BRe_flag = False
            if(HA_flag == True): # ['PT', 'INR', 'PT-%','PT%','APTT', 'FIB', 'TT', 'D-Dimer', 'FDP']
                if (HA_dic['PT'] !=None and HA_dic['INR'] != None and HA_dic['APTT'] != None and HA_dic['FIB'] != None and HA_dic['TT'] != None and HA_dic['D-Dimer'] != None):
                    if HA_dic['FDP'] == None:
                        HA_dic['FDP'] = '0'                                   
                writeFile(output_file['HA_file'],output_wb['HA_wb'], output_sheet['HA_sheet'], **HA_dic)
                HA_flag = False
            if(PCT_flag == True):
                writeFile(output_file['PCT_file'],output_wb['PCT_wb'], output_sheet['PCT_sheet'], **PCT_dic)
                PCT_flag = False
            if(CRP_flag == True):
                writeFile(output_file['CRP_file'],output_wb['CRP_wb'], output_sheet['CRP_sheet'], **CRP_dic)
                CRP_flag = False

            write_flag = False
            initDic()

        # 取出新检验号的第一行数据
        if (jianyanhao_cell_value != jianyanhao_cell_value_old):
            xiangmudaima_cell = sheet.cell(row=rows_no, column=xiangmudaima_index)
            xiangmudaima_cell_value = str(xiangmudaima_cell.value)

            # 项目名称，作为降钙素原和C反应蛋白判断的联合项
            xiangmumingcheng_cell = sheet.cell(row=rows_no, column=xiangmumingcheng_index)
            xiangmumingcheng_cell_value = str(xiangmumingcheng_cell.value)

            # 取出检验结果
            if(BRe_xmdm.count(xiangmudaima_cell_value) != 0):
                BRe_flag = True
                BRe_dic[xiangmudaima_cell_value] = str(sheet.cell(row=rows_no, column=jianyanjieguo_index).value)
                if((BRe_dic['HOSP_ID']) == None) :
                    BRe_dic['HOSP_ID'] = str(sheet.cell(row=rows_no, column=zhuyuanhao_index).value)
                    BRe_dic['TEST_NUM'] = str(sheet.cell(row=rows_no, column=jianyanhao_index).value)
                    BRe_dic['SEX'] = str(sheet.cell(row=rows_no, column=sex_index).value)
                    BRe_dic['TEST_TIME'] = str(sheet.cell(row=rows_no, column=jianyanshijian_index).value)
            elif(HA_xmdm.count(xiangmudaima_cell_value) != 0):
                HA_flag = True
                if xiangmudaima_cell_value =='PT%':
                    xiangmudaima_cell_value = 'PT-%'
                if xiangmudaima_cell_value == 'Fib':
                    xiangmudaima_cell_value = 'FIB'
                HA_dic[xiangmudaima_cell_value] = str(sheet.cell(row=rows_no, column=jianyanjieguo_index).value)
                if((HA_dic['HOSP_ID']) == None) :
                    HA_dic['HOSP_ID'] = str(sheet.cell(row=rows_no, column=zhuyuanhao_index).value)
                    HA_dic['TEST_NUM'] = str(sheet.cell(row=rows_no, column=jianyanhao_index).value)
                    HA_dic['SEX'] = str(sheet.cell(row=rows_no, column=sex_index).value)
                    HA_dic['TEST_TIME'] = str(sheet.cell(row=rows_no, column=jianyanshijian_index).value)
            # elif(xiangmudaima_cell_value in PCT_xmdm) and (PCT_xmmc in xiangmumingcheng_cell_value):
            elif (xiangmudaima_cell_value in PCT_xmdm):
                xiangmumingcheng_cell = sheet.cell(row=rows_no, column=xiangmumingcheng_index)
                xiangmumingcheng_cell_value = str(xiangmumingcheng_cell.value)
                if (PCT_xmmc in xiangmumingcheng_cell_value):
                    # 更新写标志位，是一个单项项目，只要有一次就表示已经可以写入
                    write_flag = True
                    PCT_flag = True
                    PCT_dic[xiangmudaima_cell_value] = str(sheet.cell(row=rows_no, column=jianyanjieguo_index).value)
                    if ((PCT_dic['HOSP_ID']) == None):
                        PCT_dic['HOSP_ID'] = str(sheet.cell(row=rows_no, column=zhuyuanhao_index).value)
                        PCT_dic['TEST_NUM'] = str(sheet.cell(row=rows_no, column=jianyanhao_index).value)
                        PCT_dic['SEX'] = str(sheet.cell(row=rows_no, column=sex_index).value)
                        PCT_dic['TEST_TIME'] = str(sheet.cell(row=rows_no, column=jianyanshijian_index).value)
                # elif(xiangmudaima_cell_value in CRP_xmdm) and (CRP_xmmc in xiangmumingcheng_cell_value):
            elif (xiangmudaima_cell_value in CRP_xmdm):
                # 项目名称，作为降钙素原和C反应蛋白判断的联合项
                xiangmumingcheng_cell = sheet.cell(row=rows_no, column=xiangmumingcheng_index)
                xiangmumingcheng_cell_value = str(xiangmumingcheng_cell.value)
                if (CRP_xmmc in xiangmumingcheng_cell_value):
                    # 更新写标志位，CRP存在同一个检验号下有hsCRP和CRP，且是一个单项项目，只要有一次就表示已经可以写入
                    write_flag = True
                    CRP_flag = True
                    if xiangmudaima_cell_value == 'hsCRP':
                        xiangmudaima_cell_value = 'CRP'
                    CRP_dic[xiangmudaima_cell_value] = str(sheet.cell(row=rows_no, column=jianyanjieguo_index).value)
                    if ((CRP_dic['HOSP_ID']) == None):
                        CRP_dic['HOSP_ID'] = str(sheet.cell(row=rows_no, column=zhuyuanhao_index).value)
                        CRP_dic['TEST_NUM'] = str(sheet.cell(row=rows_no, column=jianyanhao_index).value)
                        CRP_dic['SEX'] = str(sheet.cell(row=rows_no, column=sex_index).value)
                        CRP_dic['TEST_TIME'] = str(sheet.cell(row=rows_no, column=jianyanshijian_index).value)
            else:
                # 其余不做处理
                pass
            '''
            # 新检验号的第一行刚好是最后一行
            if(rows_no  == rows):
                if (BRe_flag == True):
                    writeFile(output_file['BRe_file'], output_wb['BRe_wb'], output_sheet['BRe_sheet'], **BRe_dic)
                    BRe_flag = False
                if (HA_flag == True):
                    writeFile(output_file['HA_file'],output_wb['HA_wb'], output_sheet['HA_sheet'], **HA_dic)
                    HA_flag = False
                if (PCT_flag == True):
                    writeFile(output_file['PCT_file'],output_wb['PCT_wb'], output_sheet['PCT_sheet'], **PCT_dic)
                    PCT_flag = False
                if (CRP_flag == True):
                    writeFile(output_file['CRP_file'],output_wb['CRP_wb'], output_sheet['CRP_sheet'], **CRP_dic)
                    CRP_flag = False
            '''
        jianyanhao_cell_value_old = jianyanhao_cell_value
        rows_no += 1
        
        # times
        if (info_dedug and ((rows_no % 5000) == 0)):
            print(rows_no)
            print('spend time (second): ',tm.clock()- begin_time)
            begin_time = tm.clock()

    # 新检验号的第一行刚好是最后一行
    if (BRe_flag == True):
        writeFile(output_file['BRe_file'], output_wb['BRe_wb'], output_sheet['BRe_sheet'], **BRe_dic)
        BRe_flag = False
    if (HA_flag == True):
        writeFile(output_file['HA_file'], output_wb['HA_wb'], output_sheet['HA_sheet'], **HA_dic)
        HA_flag = False
    if (PCT_flag == True):
        writeFile(output_file['PCT_file'], output_wb['PCT_wb'], output_sheet['PCT_sheet'], **PCT_dic)
        PCT_flag = False
    if (CRP_flag == True):
        writeFile(output_file['CRP_file'], output_wb['CRP_wb'], output_sheet['CRP_sheet'], **CRP_dic)
        CRP_flag = False

    print(rows_no)


if __name__ == '__main__':

    main_time = tm.clock()
 
    # 当前支持一个文件
    #file_name = 'XPY01.xlsx'
     
     # 命名行
    parser = argparse.ArgumentParser()
    parser.add_argument('-i', default="None", help='input file')
    args = parser.parse_args()
    
    file_name = args.i 
    if file_name == 'None':
         print('缺少参数：文件名！eg: dataPre_LisItem_Segmentation -i file')
         sys.exit()
      
    file_dir = os.path.dirname(os.path.realpath(__file__))
 
    if sys.platform == 'linux':
        input_file = file_dir + r'/' + file_name
    else:
#         file_dir = file_dir.rstrip(r'\venv')
#         input_file = file_dir + str(r'\?').rstrip(r'?')+ file_name
        input_file = os.path.join(os.path.abspath('..'), 'data', 'BC',file_name)
    file_path = input_file.rstrip('.xlsx')
 
    print(sys.platform)
    print(file_name)
    print(file_dir)
    print(file_path)
    print(input_file)
 
    initDic()
    initOutFile(file_path)
 
    # 调试
    info_dedug = True
    error_debug = False

    getLisInf(input_file)
    

    # 保存文件
    for key in output_file.keys():
        key_wb = key.rstrip('file') + ('wb')
        wb = output_wb[key_wb]
        wb.save(output_file[key])

    print('spend total time (second): ', tm.clock() - main_time)
