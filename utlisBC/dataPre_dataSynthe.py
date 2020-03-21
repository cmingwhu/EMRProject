'''
Created on 2019年6月6日

数据合成：取血常规结果和血培养结果合在一起
合成规则：取血培养检验时间之前3天之内最近的血常规结果 
                 如果之前3天内没有血常规结果则取之后1天的血常规结果
                 如果都没有则放弃该样本
@author: cmingwhu
'''

from openpyxl import load_workbook ,Workbook
import os,sys
import time as tm
import numpy as np
import datetime,re
import argparse




# 血培养
BCu_column = {'HOSP_ID':None,'TEST_NUM':None,'TEST_TIME':None,'SEX':None,'AGE':None,'FLAG':None}
# 血常规
BRe_column = {'HOSP_ID': None, 'TEST_NUM': None,'TEST_TIME': None, 'SEX': None, 
           'WBC': None, 'RBC': None, 'Hb': None, 'PLT': None, 'Neut%': None, 'Lymph%': None,
           'Mono%': None,  'Eos%': None, 'Baso%': None, 'Neut#': None, 'Lymph#': None, 'Mono#': None, 'Eos#': None,
           'Baso#': None, 'Hct': None, 'MCV': None, 'MCH': None, 'MCHC': None, 'RDW': None, 'MPV': None, 'Pct': None,
           'PDW': None, 'NRBC': None, 'NRBC%': None}
# PCT
PCT_column = {'HOSP_ID': None, 'TEST_NUM': None,'TEST_TIME': None, 'SEX': None, 'PCT':None}

# CRP
CRP_column = {'HOSP_ID': None, 'TEST_NUM': None,'TEST_TIME': None, 'SEX': None, 'CRP':None}

# CRP
HA_column = {'HOSP_ID': None, 'TEST_NUM': None,'TEST_TIME': None, 'SEX': None, 
            'PT':None, 'PT-%':None, 'INR':None, 'APTT':None, 'FIB':None, 'TT':None, 'D-Dimer':None, 'FDP':None}

result_column = {'HOSP_ID':None,'BCuTime':None,'BReTime':None,'SEX':None,'AGE':None,'BCuResult':None,
              'WBC': None, 'RBC': None, 'Hb': None, 'PLT': None, 'Neut%': None, 'Lymph%': None,
           'Mono%': None,  'Eos%': None, 'Baso%': None, 'Neut#': None, 'Lymph#': None, 'Mono#': None, 'Eos#': None,
           'Baso#': None, 'Hct': None, 'MCV': None, 'MCH': None, 'MCHC': None, 'RDW': None, 'MPV': None, 'Pct': None,
           'PDW': None, 'NRBC': None, 'NRBC%': None,'PCT': None,'CRP': None,'PT': None,'PT-%': None,'INR': None,'APTT': None,'FIB': None,'TT': None,'D-Dimer': None,'FDP': None}

result_dic = {'HOSP_ID':None,'BCuTime':None,'BReTime':None,'PCTTime':None,'CRPTime':None,'HATime':None,'SEX':None,'AGE':None,
              'WBC': None, 'RBC': None, 'Hb': None, 'PLT': None, 'Neut%': None, 'Lymph%': None,
           'Mono%': None,  'Eos%': None, 'Baso%': None, 'Neut#': None, 'Lymph#': None, 'Mono#': None, 'Eos#': None,
           'Baso#': None, 'Hct': None, 'MCV': None, 'MCH': None, 'MCHC': None, 'RDW': None, 'MPV': None, 'Pct': None,
           'PDW': None, 'NRBC': None, 'NRBC%': None,'PCT': None,'CRP': None,'PT': None,'PT-%': None,'INR': None,'APTT': None,'FIB': None,'TT': None,'D-Dimer': None,'FDP': None,'BCuResult':None,}

result_firRow = ['HOSP_ID','BCuTime','BReTime','PCTTime','CRPTime','HATime','SEX','AGE','WBC', 'RBC', 'Hb', 'PLT', 'Neut%', 'Lymph%',
                 'Mono%',  'Eos%', 'Baso%', 'Neut#', 'Lymph#', 'Mono#', 'Eos#','Baso#', 'Hct', 'MCV', 'MCH', 'MCHC',
                 'RDW', 'MPV', 'Pct', 'PDW', 'NRBC', 'NRBC%','PCT','CRP','PT','PT-%','INR','APTT','FIB','TT','D-Dimer','FDP','BCuResult']


result_row =[]

error_debug = False
info_dedug = False

output_file = {'Result_file': None, 'error_file': None}
output_wb = {'Result_wb': None, 'error_wb': None}
output_sheet = {'Result_sheet': None, 'error_sheet': None}

def initDic():
    for key in result_dic.keys():
        result_dic[key] = None

def initOutFile(path):
    tm.sleep(2)
    cur_time = (tm.strftime('%Y-%m-%d %H:%M:%S', tm.localtime(tm.time()))).replace(':', '-', 100)

    output_file['Result_file'] = path + '--Rusult ' + '(' + cur_time + ')' + '.xlsx'
    output_file['error_file'] = path + '--error ' + '(' + cur_time + ')' + '.xlsx'

    # 添加第一行
    for key in result_dic.keys():
        result_dic[key] = key
        
        
    for key in output_file.keys():
        file = output_file[key]
        if os.path.exists(file):
            os.remove(output_file[key])
            print('file is the same,return\n')
            exit(f0000001)
        else:
            wb = Workbook()
            sheet = wb.active          
            sheet.append(result_firRow)   
            wb.save(file)

    for key in output_wb.keys():
        file = output_file[key.rstrip('wb') + ('file')]
        key_sheet = key.rstrip('wb') + ('sheet')
        output_wb[key] = load_workbook(file)
        output_sheet[key_sheet] = output_wb[key].get_active_sheet()

    
'''
'''
def writeFile(wr_file, wb, sheet, **dic):
    row = []
    err_flag = False

    for key,value in dic.items():
        if(value == None):
            if(error_debug == False):
                return
            #print('非法的数据 %s \n'%key)
            err_flag = True
        row.append(value)
    result_row.append(row)

    if(err_flag == True):
        wr_excel = output_file['error_file']
        wb = output_wb['error_wb']
        sheet = output_sheet['error_sheet']
    else:
        wr_excel = wr_file

    sheet.append(row)
#     wb.save(wr_excel) # 浪费时间的语句

'''
数据整合 血培养和血常规
 '''
def dataInter_BCu_BRe(fileBCu, fileBRe):
    # get sheet of blood culture 
    csvBCu = load_workbook(fileBCu)
    # 获取sheet：
    sheetBCu = csvBCu.get_active_sheet()
    #  载入BRe
    csvBRe = load_workbook(fileBRe)
    sheetBRe = csvBRe.get_active_sheet()

    
    # 取血培养表所在的index
    tempList  = list(sheetBCu.rows)
    for key in BCu_column.keys():
        column_no = 1
        for cell in tempList[0]:
            if(cell.value == key):
                BCu_column[key] = column_no
                break
            column_no += 1
            
    # 取血常规表所在的index
    tempList  = list(sheetBRe.rows)
    for key in BRe_column.keys():
        column_no = 1
        for cell in tempList[0]:
            if(cell.value == key):
                BRe_column[key] = column_no
                break
            column_no += 1
            


    # 写标志位        
    write_flag = False   
    
    # 获取检验号一列的数据
    rowsBCu_no = 2     
    RT_list = ['HOSP_ID', 'TEST_NUM','TEST_TIME', 'SEX']

    BCuPatID_old = 'None'
    BCuTime_old = 'None'
    
    if (info_dedug):
        begin_time = tm.clock()
    
    # 遍历 血培养
    for cell in list(sheetBCu.columns)[1]:     
        # 行为空 结束后
        if (str(sheetBCu.cell(row=rowsBCu_no, column=BCu_column['HOSP_ID']).value) == 'None'):
            break
        # 血培养 血常规 相距时间3天
        timeT = 4
        
        for key in BCu_column.keys():
            # 获取血培养数据
            if (key == 'HOSP_ID'):
                result_dic['HOSP_ID'] = str(sheetBCu.cell(row=rowsBCu_no, column=BCu_column[key]).value)
            if (key == 'SEX'):
                result_dic['SEX'] = str(sheetBCu.cell(row=rowsBCu_no, column=BCu_column[key]).value)  
            if (key == 'AGE'):
                result_dic['AGE'] = re.sub("\D", "", str(sheetBCu.cell(row=rowsBCu_no, column=BCu_column[key]).value))             
            if (key == 'FLAG'):
                result_dic['BCuResult'] = str(sheetBCu.cell(row=rowsBCu_no, column=BCu_column[key]).value)             
            if (key == 'TEST_TIME'):
                result_dic['BCuTime'] = str(sheetBCu.cell(row=rowsBCu_no, column=BCu_column[key]).value)           
        
        # 血培养TEST_TIME
        BCuTime = result_dic['BCuTime']
        BCuPatID = result_dic['HOSP_ID']  
        
        ####   血常规   ###################         
        # 临时存行号 
        rows_no = 2
        rowstemp_no = None
        BReTime_old1 = None
        # 遍历血常规
        for cell in list(sheetBRe.columns)[1]:
            BReTime = str(sheetBRe.cell(row=rows_no, column=BRe_column['TEST_TIME']).value)
            BRePatID = str(sheetBRe.cell(row=rows_no, column=BRe_column['HOSP_ID']).value)
                                   
            # 有匹配的住院号
            if(BCuPatID == BRePatID):  
                write_flag = True       
                # 只取血培养后2天前 的血常规结果  
                BCuTime_day = datetime.datetime.strptime(BCuTime, '%Y%m%d')
                BReTime_day = datetime.datetime.strptime(BReTime, '%Y%m%d')
                timeD = (BCuTime_day- BReTime_day).days   
               
                # 如果同一天做了两次血常规 且是时间最短的 目前只取了第一次时间最短的              
                if ((abs(timeD) < timeT) and (timeD >-3)):
                    rowstemp_no = rows_no 
                    timeT = abs(timeD)
#                     print('血常规: '+BRePatID +' ' +BReTime)
                    BReTime_old1 = BReTime                        
            rows_no += 1
        
        #####   write file   ###################   
        if (rowstemp_no != None):
            if(BCuPatID_old != BCuPatID):
                for key in BRe_column.keys():
                        if(key not in RT_list):
                            result_dic[key] = str(sheetBRe.cell(row=rowstemp_no, column=BRe_column[key]).value)  
                # 血培养时间
                result_dic['BReTime'] = str(sheetBRe.cell(row=rowstemp_no, column=BRe_column['TEST_TIME']).value)  
                writeFile(output_file['Result_file'], output_wb['Result_wb'], output_sheet['Result_sheet'],**result_dic)     
                BCuPatID_old = BCuPatID
                BReTime_old = result_dic['BReTime']
            else:
                # 相同住院号多次血培养 对应相同的血常规结果 只存1次
                if (str(sheetBRe.cell(row=rowstemp_no, column=BRe_column['TEST_TIME']).value)  != BReTime_old):
                    for key in BRe_column.keys():
                        if(key not in RT_list):
                            result_dic[key] = str(sheetBRe.cell(row=rowstemp_no, column=BRe_column[key]).value)  
                            # 血培养时间
                    result_dic['BReTime'] = str(sheetBRe.cell(row=rowstemp_no, column=BRe_column['TEST_TIME']).value)  
                    writeFile(output_file['Result_file'], output_wb['Result_wb'], output_sheet['Result_sheet'],**result_dic)     
                    BCuPatID_old = BCuPatID
                    BReTime_old = result_dic['BReTime']

        rowsBCu_no += 1
        
        # times
        if (info_dedug and ((rowsBCu_no % 100) == 0)):
            print(rowsBCu_no)
            print('spend time (second): ',tm.clock()- begin_time)
            begin_time = tm.clock()

    print ('OVER!')



'''
所有特征整合 阳性患者
'''
def dataPosInter_All(fileBCu, fileBRe, filePCT, fileCRP, fileHA):
    # get sheet of blood culture 
    csvBCu = load_workbook(fileBCu)
    # 获取sheet：
    sheetBCu = csvBCu.get_active_sheet()
    #  载入BRe
    csvBRe = load_workbook(fileBRe)
    sheetBRe = csvBRe.get_active_sheet()
    # 载入PCT
    csvPCT = load_workbook(filePCT)
    sheetPCT = csvPCT.get_active_sheet()
    # 载入CRP
    csvCRP = load_workbook(fileCRP)
    sheetCRP= csvCRP.get_active_sheet()
    # 载入HA
    csvHA = load_workbook(fileHA)
    sheetHA= csvHA.get_active_sheet()
    
    # 取血培养表所在的index
    tempList  = list(sheetBCu.rows)
    for key in BCu_column.keys():
        column_no = 1
        for cell in tempList[0]:
            if(cell.value == key):
                BCu_column[key] = column_no
                break
            column_no += 1
            
    # 取血常规表所在的index
    tempList  = list(sheetBRe.rows)
    for key in BRe_column.keys():
        column_no = 1
        for cell in tempList[0]:
            if(cell.value == key):
                BRe_column[key] = column_no
                break
            column_no += 1
            
    # 取PCT表所在的index
    tempList  = list(sheetPCT.rows)
    for key in PCT_column.keys():
        column_no = 1
        for cell in tempList[0]:
            if(cell.value == key):
                PCT_column[key] = column_no
                break
            column_no += 1
     
    # 取CRP表所在的index
    tempList  = list(sheetCRP.rows)
    for key in CRP_column.keys():
        column_no = 1
        for cell in tempList[0]:
            if(cell.value == key):
                CRP_column[key] = column_no
                break
            column_no += 1
  
    # 取HA表所在的index
    tempList  = list(sheetHA.rows)
    for key in HA_column.keys():
        column_no = 1
        for cell in tempList[0]:
            if(cell.value == key):
                HA_column[key] = column_no
                break
            column_no += 1

    
    # 获取检验号一列的数据
    rowsBCu_no = 2     
    RT_list = ['HOSP_ID', 'TEST_NUM','TEST_TIME', 'SEX']

    BCuPatID_old = 'None'
    BCuTime_old = 'None'
    
    if (info_dedug):
        begin_time = tm.clock()
    
    # 遍历 血培养
    for cell in list(sheetBCu.columns)[1]:     
        # 行为空 结束后
        if (str(sheetBCu.cell(row=rowsBCu_no, column=BCu_column['HOSP_ID']).value) == 'None'):
            break
        # 血培养 血常规 相距时间3天
        timeT = 4
        
#         if (str(sheetBCu.cell(row=rowsBCu_no, column=BCu_column['HOSP_ID']).value) == 'ZY010002995658'):
#             print()
        
        for key in BCu_column.keys():
            # 获取血培养数据
            if (key == 'HOSP_ID'):
                result_dic['HOSP_ID'] = str(sheetBCu.cell(row=rowsBCu_no, column=BCu_column[key]).value)
            if (key == 'SEX'):
                result_dic['SEX'] = str(sheetBCu.cell(row=rowsBCu_no, column=BCu_column[key]).value)  
            if (key == 'AGE'):
                result_dic['AGE'] = re.sub("\D", "", str(sheetBCu.cell(row=rowsBCu_no, column=BCu_column[key]).value))             
            if (key == 'FLAG'):
                result_dic['BCuResult'] = str(sheetBCu.cell(row=rowsBCu_no, column=BCu_column[key]).value)             
            if (key == 'TEST_TIME'):
                result_dic['BCuTime'] = str(sheetBCu.cell(row=rowsBCu_no, column=BCu_column[key]).value)           
        
        # 血培养TEST_TIME
        BCuTime = result_dic['BCuTime']
        BCuPatID = result_dic['HOSP_ID']  
        
        ####   血常规   ###################         
        # 临时存行号 
        rows_no = 2
        rowstemp_no = None
        # 遍历血常规
        for cell in list(sheetBRe.columns)[1]:
            BReTime = str(sheetBRe.cell(row=rows_no, column=BRe_column['TEST_TIME']).value)
            BRePatID = str(sheetBRe.cell(row=rows_no, column=BRe_column['HOSP_ID']).value)
                                   
            # 有匹配的住院号
            if(BCuPatID == BRePatID):  
                write_flag = True       
                # 只取血培养后2天前 的血常规结果  
                BCuTime_day = datetime.datetime.strptime(BCuTime, '%Y%m%d')
                BReTime_day = datetime.datetime.strptime(BReTime, '%Y%m%d')
                timeD = (BCuTime_day- BReTime_day).days   
               
                # 如果同一天做了两次血常规 且是时间最短的 目前只取了第一次时间最短的              
                if ((abs(timeD) < timeT) and (timeD >-2)):
                    rowstemp_no = rows_no 
                    timeT = abs(timeD)                      
            rows_no += 1
     
            
        #####   PCT  ###################           
        PCT_res = 0
        rows_no = 2
        rowsPCT_no = None
        # 遍历PCT
        for cell in list(sheetPCT.columns)[1]:
            PCTTime = str(sheetPCT.cell(row=rows_no, column=PCT_column['TEST_TIME']).value)
            PCTPatID = str(sheetPCT.cell(row=rows_no, column=PCT_column['HOSP_ID']).value)
                                    
            # 有匹配的住院号
            if(BCuPatID == PCTPatID):        
                # 取血培养之前的，最高的PCT值
                PCTRse_ = str(sheetPCT.cell(row=rows_no, column=PCT_column['PCT']).value)
                if ('>' in PCTRse_):
                    if ((float(PCTRse_[1:])*3) > PCT_res):
                        rowsPCT_no = rows_no
                        PCT_res = float(PCTRse_[1:])*3
                elif('<'in PCTRse_):
                    pass
                else:
                    if float(PCTRse_) > PCT_res:
                        rowsPCT_no = rows_no
                        PCT_res = float(PCTRse_)                          
            rows_no += 1
             
        ####   CRP  ###################           
        CRP_res = 0.0
        rows_no = 2
        rowsCRP_no = None
        # 遍历CRP
        for cell in list(sheetCRP.columns)[1]:
            CRPTime = str(sheetCRP.cell(row=rows_no, column=CRP_column['TEST_TIME']).value)
            CRPPatID = str(sheetCRP.cell(row=rows_no, column=CRP_column['HOSP_ID']).value)
            
                                    
            # 有匹配的住院号
            if(BCuPatID == CRPPatID):  
                # 取血培养之前的，最高的CRP值
                CRPRse_ = str(sheetCRP.cell(row=rows_no, column=CRP_column['CRP']).value)
                if ('>' in CRPRse_):
                    if ((float(CRPRse_[1:])*3) > CRP_res):
                        rowsCRP_no = rows_no
                        CRP_res = float(CRPRse_[1:])*3
                elif('<'in CRPRse_):
                    pass
                else:
                    if CRPRse_[-1:] == '.':
                        if float(CRPRse_[0:-1]) < CRP_res:
                            rowsCRP_no = rows_no
                            CRP_res = float(CRPRse_)  
                    else:                        
                        if float(CRPRse_) < CRP_res:
                            rowsCRP_no = rows_no
                            CRP_res = float(CRPRse_)                 
            rows_no += 1
             
             
        ####   HA  ###################           
        HA_res = 0
        rows_no = 2
        rowsHA_no = None
        # 遍历HA
        for cell in list(sheetHA.columns)[1]:
            HATime = str(sheetHA.cell(row=rows_no, column=HA_column['TEST_TIME']).value)
            HAPatID = str(sheetHA.cell(row=rows_no, column=HA_column['HOSP_ID']).value)
                                    
            # 有匹配的住院号
            if(BCuPatID == HAPatID):  
                # 取血培养之前的，最高的HA值
                HARse_ = str(sheetHA.cell(row=rows_no, column=HA_column['PT']).value)
                if HARse_ == '不凝':
                    HARse_ = 0.0
                if (float(HARse_) > HA_res):
                    rowsHA_no = rows_no
                    HA_res = float(HARse_)                
            rows_no += 1
            
           
        #####   write file   ###################   
        if (rowstemp_no != None):
            #  血常规
            for key in BRe_column.keys():
                    if(key not in RT_list):
                        result_dic[key] = str(sheetBRe.cell(row=rowstemp_no, column=BRe_column[key]).value)  
            # 血培养时间
            result_dic['BReTime'] = str(sheetBRe.cell(row=rowstemp_no, column=BRe_column['TEST_TIME']).value)                
            
        # PCT
        if (rowsPCT_no != None):
            for key in PCT_column.keys():
                if(key not in RT_list):
                    if('>' in str(sheetPCT.cell(row=rowsPCT_no, column=PCT_column[key]).value) ):
                        result_dic[key] = str(float(str(sheetPCT.cell(row=rowsPCT_no, column=PCT_column[key]).value)[1:])*3)
                    elif('<' in PCTRse_):
                        result_dic[key] = '0'
                    else:
                        result_dic[key] = str(sheetPCT.cell(row=rowsPCT_no, column=PCT_column[key]).value) 
            # PCT时间
            result_dic['PCTTime'] = str(sheetPCT.cell(row=rowsPCT_no, column=PCT_column['TEST_TIME']).value)
        else:
            for key in PCT_column.keys():
                if(key not in RT_list):
                    result_dic[key] = '15.6'
            # PCT时间
            result_dic['PCTTime'] = BCuTime
            
        # CRP
        if (rowsCRP_no != None):
            for key in CRP_column.keys():
                if(key not in RT_list):
                    if('>' in str(sheetCRP.cell(row=rowsCRP_no, column=CRP_column[key]).value) ):
                        result_dic[key] = str(float(str(sheetCRP.cell(row=rowsCRP_no, column=CRP_column[key]).value)[1:])*3)
                    elif('<' in PCTRse_):
                        result_dic[key] = '0'
                    else:
                        result_dic[key] = str(sheetCRP.cell(row=rowsCRP_no, column=CRP_column[key]).value) 
            # CRP时间
            result_dic['CRPTime'] = str(sheetCRP.cell(row=rowsCRP_no, column=CRP_column['TEST_TIME']).value)
        else:
            for key in CRP_column.keys():
                if(key not in RT_list):
                    result_dic[key] = '15.6'
            # PCT时间
            result_dic['CRPTime'] = BCuTime
            
        # HA
        if (rowsHA_no != None):
            for key in HA_column.keys():
                if(key not in RT_list):
                    if str(sheetHA.cell(row=rowsHA_no, column=HA_column[key]).value) =='不凝':
                        result_dic[key] = '0.0'
                    else:
                        result_dic[key] = str(sheetHA.cell(row=rowsHA_no, column=HA_column[key]).value) 
            # HA时间
            result_dic['HATime'] = str(sheetHA.cell(row=rowsHA_no, column=HA_column['TEST_TIME']).value)
        else:
            for key in HA_column.keys():
                if(key not in RT_list):
                    result_dic[key] = '15.6'
            # PCT时间
            result_dic['HATime'] = BCuTime    
            
        writeFile(output_file['Result_file'], output_wb['Result_wb'], output_sheet['Result_sheet'],**result_dic)  
        
        rowsBCu_no += 1
        
#         print(BCuPatID)
        # times
        if (info_dedug and ((rowsBCu_no % 100) == 0)):
            print(rowsBCu_no)
            print('spend time (second): ',tm.clock()- begin_time)
            begin_time = tm.clock()
            
    print ('OVER!')   
    
    
'''
所有特征整合 阴性患者
'''
def dataNegInter_All(fileBCu, fileBRe, filePCT, fileCRP, fileHA):
    # get sheet of blood culture 
    csvBCu = load_workbook(fileBCu)
    # 获取sheet：
    sheetBCu = csvBCu.get_active_sheet()
    #  载入BRe
    csvBRe = load_workbook(fileBRe)
    sheetBRe = csvBRe.get_active_sheet()
    # 载入PCT
    csvPCT = load_workbook(filePCT)
    sheetPCT = csvPCT.get_active_sheet()
    # 载入CRP
    csvCRP = load_workbook(fileCRP)
    sheetCRP= csvCRP.get_active_sheet()
    # 载入HA
    csvHA = load_workbook(fileHA)
    sheetHA= csvHA.get_active_sheet()
    
    # 取血培养表所在的index
    tempList  = list(sheetBCu.rows)
    for key in BCu_column.keys():
        column_no = 1
        for cell in tempList[0]:
            if(cell.value == key):
                BCu_column[key] = column_no
                break
            column_no += 1
            
    # 取血常规表所在的index
    tempList  = list(sheetBRe.rows)
    for key in BRe_column.keys():
        column_no = 1
        for cell in tempList[0]:
            if(cell.value == key):
                BRe_column[key] = column_no
                break
            column_no += 1
            
    # 取PCT表所在的index
    tempList  = list(sheetPCT.rows)
    for key in PCT_column.keys():
        column_no = 1
        for cell in tempList[0]:
            if(cell.value == key):
                PCT_column[key] = column_no
                break
            column_no += 1
     
    # 取CRP表所在的index
    tempList  = list(sheetCRP.rows)
    for key in CRP_column.keys():
        column_no = 1
        for cell in tempList[0]:
            if(cell.value == key):
                CRP_column[key] = column_no
                break
            column_no += 1
  
    # 取HA表所在的index
    tempList  = list(sheetHA.rows)
    for key in HA_column.keys():
        column_no = 1
        for cell in tempList[0]:
            if(cell.value == key):
                HA_column[key] = column_no
                break
            column_no += 1

    
    # 获取检验号一列的数据
    rowsBCu_no = 2     
    RT_list = ['HOSP_ID', 'TEST_NUM','TEST_TIME', 'SEX']

    BCuPatID_old = 'None'
    BCuTime_old = 'None'
    
    if (info_dedug):
        begin_time = tm.clock()
    
    # 遍历 血培养
    for cell in list(sheetBCu.columns)[1]:     
        # 行为空 结束后
        if (str(sheetBCu.cell(row=rowsBCu_no, column=BCu_column['HOSP_ID']).value) == 'None'):
            break
        # 血培养 血常规 相距时间3天
        timeT = 4
        
        
        for key in BCu_column.keys():
            # 获取血培养数据
            if (key == 'HOSP_ID'):
                result_dic['HOSP_ID'] = str(sheetBCu.cell(row=rowsBCu_no, column=BCu_column[key]).value)
            if (key == 'SEX'):
                result_dic['SEX'] = str(sheetBCu.cell(row=rowsBCu_no, column=BCu_column[key]).value)  
            if (key == 'AGE'):
                result_dic['AGE'] = re.sub("\D", "", str(sheetBCu.cell(row=rowsBCu_no, column=BCu_column[key]).value))             
            if (key == 'FLAG'):
                result_dic['BCuResult'] = str(sheetBCu.cell(row=rowsBCu_no, column=BCu_column[key]).value)             
            if (key == 'TEST_TIME'):
                result_dic['BCuTime'] = str(sheetBCu.cell(row=rowsBCu_no, column=BCu_column[key]).value)           
        
        # 血培养TEST_TIME
        BCuTime = result_dic['BCuTime']
        BCuPatID = result_dic['HOSP_ID']  
        
        ####   血常规   ###################         
        # 临时存行号 
        BCu_res = 6.50
        rows_no = 2
        rowstemp_no = None
        # 遍历血常规
        for cell in list(sheetBRe.columns)[1]:
            BReTime = str(sheetBRe.cell(row=rows_no, column=BRe_column['TEST_TIME']).value)
            BRePatID = str(sheetBRe.cell(row=rows_no, column=BRe_column['HOSP_ID']).value)
                                   
            # 有匹配的住院号
            if(BCuPatID == BRePatID):  
                # 取入院后第一次血常规
                rowstemp_no = rows_no
                break
                 
            rows_no += 1
     
            
        #####   PCT  ###################           
        PCT_res = 10.0
        rows_no = 2
        rowsPCT_no = None
        # 遍历PCT
        for cell in list(sheetPCT.columns)[1]:
            PCTTime = str(sheetPCT.cell(row=rows_no, column=PCT_column['TEST_TIME']).value)
            PCTPatID = str(sheetPCT.cell(row=rows_no, column=PCT_column['HOSP_ID']).value)
                                    
            # 有匹配的住院号
            if(BCuPatID == PCTPatID):        
                # 取血培养之前的，最低的PCT值
                PCTRse_ = str(sheetPCT.cell(row=rows_no, column=PCT_column['PCT']).value)
                if ('>' in PCTRse_):
                    if ((float(PCTRse_[1:])*3) < PCT_res):
                        rowsPCT_no = rows_no
                        PCT_res = float(PCTRse_[1:])*3
                elif('<'in PCTRse_):
                    pass
                else:
                    if float(PCTRse_) < PCT_res:
                        rowsPCT_no = rows_no
                        PCT_res = float(PCTRse_)                          
            rows_no += 1
             
        ####   CRP  ###################           
        CRP_res = 10.0
        rows_no = 2
        rowsCRP_no = None
        # 遍历CRP
        for cell in list(sheetCRP.columns)[1]:
            CRPTime = str(sheetCRP.cell(row=rows_no, column=CRP_column['TEST_TIME']).value)
            CRPPatID = str(sheetCRP.cell(row=rows_no, column=CRP_column['HOSP_ID']).value)
            
                                    
            # 有匹配的住院号
            if(BCuPatID == CRPPatID):  
                # 取血培养之前的，最低的CRP值
                CRPRse_ = str(sheetCRP.cell(row=rows_no, column=CRP_column['CRP']).value)
                if ('>' in CRPRse_):
                    if ((float(CRPRse_[1:])*3) < CRP_res):
                        rowsCRP_no = rows_no
                        CRP_res = float(CRPRse_[1:])*3
                elif('<'in CRPRse_):
                    pass
                else:
                    if CRPRse_[-1:] == '.':
                        if float(CRPRse_[0:-1]) < CRP_res:
                            rowsCRP_no = rows_no
                            CRP_res = float(CRPRse_)  
                    else:                        
                        if float(CRPRse_) < CRP_res:
                            rowsCRP_no = rows_no
                            CRP_res = float(CRPRse_)                 
            rows_no += 1
             
             
        ####   HA  ###################           
        HA_res = 10.0
        rows_no = 2
        rowsHA_no = None
        # 遍历HA
        for cell in list(sheetHA.columns)[1]:
            HATime = str(sheetHA.cell(row=rows_no, column=HA_column['TEST_TIME']).value)
            HAPatID = str(sheetHA.cell(row=rows_no, column=HA_column['HOSP_ID']).value)
                                    
            # 有匹配的住院号
            if(BCuPatID == HAPatID):  
                # 取第一次血凝结果
                HARse_ = str(sheetHA.cell(row=rows_no, column=HA_column['PT']).value)
                if HARse_ == '不凝':
                    HARse_ = 0.0
                    
                rowsHA_no = rows_no

                             
            rows_no += 1
            
           
        #####   write file   ###################   
        if (rowstemp_no != None):
            #  血常规
            for key in BRe_column.keys():
                    if(key not in RT_list):
                        result_dic[key] = str(sheetBRe.cell(row=rowstemp_no, column=BRe_column[key]).value)  
            # 血培养时间
            result_dic['BReTime'] = str(sheetBRe.cell(row=rowstemp_no, column=BRe_column['TEST_TIME']).value)                
            
        # PCT
        if (rowsPCT_no != None):
            for key in PCT_column.keys():
                if(key not in RT_list):
                    if('>' in str(sheetPCT.cell(row=rowsPCT_no, column=PCT_column[key]).value) ):
                        result_dic[key] = str(float(str(sheetPCT.cell(row=rowsPCT_no, column=PCT_column[key]).value)[1:])*3)
                    elif('<' in PCTRse_):
                        result_dic[key] = '0'
                    else:
                        result_dic[key] = str(sheetPCT.cell(row=rowsPCT_no, column=PCT_column[key]).value) 
            # PCT时间
            result_dic['PCTTime'] = str(sheetPCT.cell(row=rowsPCT_no, column=PCT_column['TEST_TIME']).value)
        else:
            for key in PCT_column.keys():
                if(key not in RT_list):
                    result_dic[key] = '0'
            # PCT时间
            result_dic['PCTTime'] = BCuTime
            
        # CRP
        if (rowsCRP_no != None):
            for key in CRP_column.keys():
                if(key not in RT_list):
                    if('>' in str(sheetCRP.cell(row=rowsCRP_no, column=CRP_column[key]).value) ):
                        result_dic[key] = str(float(str(sheetCRP.cell(row=rowsCRP_no, column=CRP_column[key]).value)[1:])*3)
                    elif('<' in PCTRse_):
                        result_dic[key] = '0'
                    else:
                        result_dic[key] = str(sheetCRP.cell(row=rowsCRP_no, column=CRP_column[key]).value) 
            # CRP时间
            result_dic['CRPTime'] = str(sheetCRP.cell(row=rowsCRP_no, column=CRP_column['TEST_TIME']).value)
        else:
            for key in CRP_column.keys():
                if(key not in RT_list):
                    result_dic[key] = '0'
            # PCT时间
            result_dic['CRPTime'] = BCuTime
            
        # HA
        if (rowsHA_no != None):
            for key in HA_column.keys():
                if(key not in RT_list):
                    if str(sheetHA.cell(row=rowsHA_no, column=HA_column[key]).value) =='不凝':
                        result_dic[key] = '0.0'
                    else:
                        result_dic[key] = str(sheetHA.cell(row=rowsHA_no, column=HA_column[key]).value) 
            # HA时间
            result_dic['HATime'] = str(sheetHA.cell(row=rowsHA_no, column=HA_column['TEST_TIME']).value)
        else:
            for key in HA_column.keys():
                if(key not in RT_list):
                    result_dic[key] = '0'
            # PCT时间
            result_dic['HATime'] = BCuTime    
            
        writeFile(output_file['Result_file'], output_wb['Result_wb'], output_sheet['Result_sheet'],**result_dic)  
        
        rowsBCu_no += 1
        
        # times
        if (info_dedug and ((rowsBCu_no % 100) == 0)):
            print(rowsBCu_no)
            print('spend time (second): ',tm.clock()- begin_time)
            begin_time = tm.clock()
            
    print ('OVER!')       


'''
 随机生成正常阴性样本
 '''
def dataRandomGen_NEG():
    RT_list = ['HOSP_ID', 'TEST_NUM','TEST_TIME']
    for i in range(3500): # 随机数个数
        for key in result_column.keys():
            if key == 'HOSP_ID':
                result_dic[key] = str(i)
            if(key not in RT_list):
                if key == 'AGE':
                    result_dic[key] = np.random.randint(18,60)
                elif key == 'SEX':
                    result_dic[key] = np.random.randint(2)
                elif key == 'WBC':
                    result_dic[key] = str(round(np.random.uniform(3.5,9.5),2))
                elif key == 'Hb':
                    result_dic[key] = str(round(np.random.uniform(115,150),2))
                elif key =='RBC':
                    result_dic[key] = str(round(np.random.uniform(3.8,5.1),2))
                elif key =='PLT':
                    result_dic[key] = str(round(np.random.uniform(125,350),2))
                elif key =='Neut%':
                    result_dic[key] = str(round(np.random.uniform(40,75),2))
                elif key =='Lymph%':
                    result_dic[key] = str(round(np.random.uniform(20,50),2))
                elif key =='Mono%':
                    result_dic[key] = str(round(np.random.uniform(3,10),2))
                elif key =='Mono#':
                    result_dic[key] = str(round(np.random.uniform(0.1,0.6),2))  
                elif key =='Eos%':
                    result_dic[key] = str(round(np.random.uniform(0,1),2))
                elif key =='Eos#':
                    result_dic[key] = str(round(np.random.uniform(0.02,0.52),2))
                elif key =='Neut#':
                    result_dic[key] = str(round(np.random.uniform(1.8,6.3),2))
                elif key =='Lymph#':
                    result_dic[key] = str(round(np.random.uniform(20,50),2))
                elif key =='Baso#':
                    result_dic[key] = str(round(np.random.uniform(0,0.06),2))
                elif key =='Baso%':
                    result_dic[key] = str(round(np.random.uniform(0,1),2))
                elif key =='Hct':
                    result_dic[key] = str(round(np.random.uniform(0.35,0.45),3))
                elif key =='MCV':
                    result_dic[key] = str(round(np.random.uniform(82,100),2))
                elif key =='MCH':
                    result_dic[key] = str(round(np.random.uniform(27,34),2))
                elif key =='MCHC':
                    result_dic[key] = str(round(np.random.uniform(316,354),2))
                elif key =='RDW':
                    result_dic[key] = str(round(np.random.uniform(11.5,14.5),2))
                elif key =='MPV':
                    result_dic[key] = str(round(np.random.uniform(6,12),2))
                elif key =='Pct':
                    result_dic[key] = str(round(np.random.uniform(0.11,0.28),2))
                elif key =='PDW':
                    result_dic[key] = str(round(np.random.uniform(9,17),2))
                elif key =='NRBC':
                    result_dic[key] ='0'
                elif key =='NRBC%':
                    result_dic[key] = '0'
                elif key =='PCT':
                    result_dic[key] = str(round(np.random.uniform(0,0.1),3))
                elif key =='CRP':
                    result_dic[key] = str(round(np.random.uniform(0,5),3))
                elif key =='PT':
                    result_dic[key] = str(round(np.random.uniform(8.8,13.6),3))
                elif key =='PT-%':
                    result_dic[key] = str(round(np.random.uniform(70,150),3))
                elif key =='INR':
                    result_dic[key] = str(round(np.random.uniform(0.8,1.6),3))
                elif key =='APTT':
                    result_dic[key] = str(round(np.random.uniform(26,40),3))
                elif key =='FIB':
                    result_dic[key] = str(round(np.random.uniform(2,4),3))
                elif key =='TT':
                    result_dic[key] = str(round(np.random.uniform(10,18),3))
                elif key =='D-Dimer':
                    result_dic[key] = str(round(np.random.uniform(0,0.3),3))
                elif key =='FDP':
                    result_dic[key] = str(round(np.random.uniform(0,5),3)) 
                                   
        writeFile(output_file['Result_file'], output_wb['Result_wb'], output_sheet['Result_sheet'],**result_dic)                 
    print ('OVER!')
    


'''
args set 
'''
def _process_args():
    parser = argparse.ArgumentParser(description='This is a temporal expression extraction sample program')
    parser.add_argument('-i', default="None", help='input BCu file')
    parser.add_argument('-r', default="None", help='input BRe file')
    parser.add_argument('-p', default="None", help='input PCT file')
    parser.add_argument('-c', default="None", help='input CRP file')

    return parser.parse_args(sys.argv[1:])

if __name__ == '__main__':
    main_time = tm.clock()
 
    
    # 命名行
    parser = argparse.ArgumentParser()
    parser.add_argument('-m',default='POS',help='POS, NEG, Random')
    parser.add_argument('-i', default="None", help='input BCu file')
    parser.add_argument('-r', default="None", help='input BRe file')
    parser.add_argument('-p', default="None", help='input PCT file')
    parser.add_argument('-c', default="None", help='input CRP file')
    parser.add_argument('-ha', default="None", help='input HA file')
    args = parser.parse_args()
    
    file_mode = args.m
    file_name = args.i 
    file_BRe = args.r
    file_PCT = args.p
    file_CRP = args.c
    file_HA = args.ha
    if file_mode =='None' or file_name == 'None' or file_BRe =='None' or file_PCT == 'None':
         print('缺少参数：文件名！eg: dataPre_LisItem_Segmentation -i BCu file -r BRe file -p PCT file')
         sys.exit()
         
    file_dir = os.path.dirname(os.path.realpath(__file__))     
 
    if sys.platform == 'linux':
        input_file = file_dir + r'/' + file_name
        BRe_file = file_dir + r'/' + file_BRe    
        PCT_file = file_dir + r'/' + file_PCT  
        CRP_file = file_dir + r'/' + file_CRP  
        HA_file = file_dir + r'/' + file_HA   
        input_file_NEG = file_dir + r'/' +'XPY_N'    
    else:
        input_file = os.path.join(os.path.abspath('..'), 'data', 'BC',file_name)     
        BRe_file = os.path.join(os.path.abspath('..'), 'data', 'BC',file_BRe)     
        PCT_file = os.path.join(os.path.abspath('..'), 'data', 'BC',file_PCT)   
        CRP_file = os.path.join(os.path.abspath('..'), 'data', 'BC',file_CRP)   
        HA_file = os.path.join(os.path.abspath('..'), 'data', 'BC',file_HA) 
        input_file_NEG = os.path.join(os.path.abspath('..'), 'data', 'BC','XPY_N')      
    file_path = input_file.rstrip('--BCu.xlsx')
 
    print(sys.platform)
    print('Model of file: '+ file_mode)
    print(file_name)
    print(file_dir)
    print(file_path)
    print(input_file)
    
     # 调试
    info_dedug = True
    error_debug = False
    
    initDic()
    
    if(file_mode == 'POS'):
        # 阳性数据关联
        initOutFile(file_path)
        dataPosInter_All(input_file,BRe_file,PCT_file,CRP_file,HA_file)
    elif(file_mode == 'NEG'):
        # 阴性数据关联
        initOutFile(file_path)
        dataNegInter_All(input_file,BRe_file,PCT_file,CRP_file,HA_file)
    elif(file_mode == 'Random'):
        # 仿真阴性数据    
        input_file_NEG = os.path.join(os.path.abspath('..'), 'data', 'BC','XPY_N')     
        initOutFile(input_file_NEG)
        dataRandomGen_NEG()
        
     
    # 仿真阴性数据    
#     input_file_NEG = os.path.join(os.path.abspath('..'), 'data', 'BC','XPY_N')     
#     initOutFile(input_file_NEG)
#     dataRandomGen_NEG()

    # 保存文件
    for key in output_file.keys():
        key_wb = key.rstrip('file') + ('wb')
        wb = output_wb[key_wb]
        wb.save(output_file[key])

    print('spend total time (second): ', tm.clock() - main_time)